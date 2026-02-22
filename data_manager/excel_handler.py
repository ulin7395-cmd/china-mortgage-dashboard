import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd

from config.constants import (
    SHEET_LOAN_PLANS, SHEET_RATE_ADJUSTMENTS, SHEET_REPAYMENT_SCHEDULE,
    SHEET_PREPAYMENTS, SHEET_CONFIG,
    LOAN_PLANS_COLUMNS, RATE_ADJUSTMENTS_COLUMNS,
    REPAYMENT_SCHEDULE_COLUMNS, PREPAYMENTS_COLUMNS, CONFIG_COLUMNS,
)
from config.settings import EXCEL_FILE, DATA_DIR, DEFAULT_COMMERCIAL_RATE, DEFAULT_PROVIDENT_RATE, DEFAULT_INFLATION_RATE


def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _default_config_rows() -> List[dict]:
    now = datetime.now().isoformat()
    return [
        {"key": "lpr_5y", "value": str(DEFAULT_COMMERCIAL_RATE), "description": "5年期以上LPR", "updated_at": now},
        {"key": "provident_rate", "value": str(DEFAULT_PROVIDENT_RATE), "description": "公积金贷款利率", "updated_at": now},
        {"key": "inflation_rate", "value": str(DEFAULT_INFLATION_RATE), "description": "年通胀率", "updated_at": now},
        {"key": "provident_limit", "value": "120", "description": "公积金贷款上限(万元)", "updated_at": now},
    ]


def init_excel(filepath: Path = EXCEL_FILE):
    """初始化 Excel 文件，创建所有 Sheet 和表头"""
    _ensure_data_dir()
    if filepath.exists():
        return

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        pd.DataFrame(columns=LOAN_PLANS_COLUMNS).to_excel(
            writer, sheet_name=SHEET_LOAN_PLANS, index=False)
        pd.DataFrame(columns=RATE_ADJUSTMENTS_COLUMNS).to_excel(
            writer, sheet_name=SHEET_RATE_ADJUSTMENTS, index=False)
        pd.DataFrame(columns=REPAYMENT_SCHEDULE_COLUMNS).to_excel(
            writer, sheet_name=SHEET_REPAYMENT_SCHEDULE, index=False)
        pd.DataFrame(columns=PREPAYMENTS_COLUMNS).to_excel(
            writer, sheet_name=SHEET_PREPAYMENTS, index=False)
        config_df = pd.DataFrame(_default_config_rows(), columns=CONFIG_COLUMNS)
        config_df.to_excel(writer, sheet_name=SHEET_CONFIG, index=False)


def backup_excel(filepath: Path = EXCEL_FILE):
    """写入前自动备份"""
    if filepath.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = filepath.with_suffix(f".xlsx.bak_{ts}")
        shutil.copy2(filepath, backup_path)
        # 只保留最近 5 个备份
        backups = sorted(filepath.parent.glob(f"{filepath.stem}.xlsx.bak_*"))
        for old in backups[:-5]:
            old.unlink()


def read_sheet(sheet_name: str, filepath: Path = EXCEL_FILE) -> pd.DataFrame:
    """读取指定 Sheet"""
    init_excel(filepath)
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl")
    except ValueError:
        df = pd.DataFrame()
    return df


def write_sheet(df: pd.DataFrame, sheet_name: str, filepath: Path = EXCEL_FILE):
    """写入指定 Sheet（覆盖该 Sheet，保留其他 Sheet）"""
    init_excel(filepath)
    backup_excel(filepath)

    from openpyxl import load_workbook
    wb = load_workbook(filepath)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.save(filepath)

    with pd.ExcelWriter(filepath, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


# ---- 贷款方案 CRUD ----

def get_all_plans(filepath: Path = EXCEL_FILE) -> pd.DataFrame:
    return read_sheet(SHEET_LOAN_PLANS, filepath)


def get_plan_by_id(plan_id: str, filepath: Path = EXCEL_FILE) -> Optional[pd.Series]:
    df = get_all_plans(filepath)
    match = df[df["plan_id"] == plan_id]
    if match.empty:
        return None
    return match.iloc[0]


def save_plan(plan_dict: dict, filepath: Path = EXCEL_FILE):
    df = get_all_plans(filepath)
    existing = df[df["plan_id"] == plan_dict["plan_id"]]
    if not existing.empty:
        for col in plan_dict:
            if col in df.columns:
                df.loc[df["plan_id"] == plan_dict["plan_id"], col] = plan_dict[col]
    else:
        new_row = pd.DataFrame([plan_dict])
        df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, SHEET_LOAN_PLANS, filepath)


def delete_plan(plan_id: str, filepath: Path = EXCEL_FILE):
    df = get_all_plans(filepath)
    df = df[df["plan_id"] != plan_id]
    write_sheet(df, SHEET_LOAN_PLANS, filepath)
    # 同时删除关联数据
    for sheet in [SHEET_REPAYMENT_SCHEDULE, SHEET_RATE_ADJUSTMENTS, SHEET_PREPAYMENTS]:
        sdf = read_sheet(sheet, filepath)
        if "plan_id" in sdf.columns:
            sdf = sdf[sdf["plan_id"] != plan_id]
            write_sheet(sdf, sheet, filepath)


# ---- 还款计划 ----

def get_repayment_schedule(plan_id: str, filepath: Path = EXCEL_FILE) -> pd.DataFrame:
    df = read_sheet(SHEET_REPAYMENT_SCHEDULE, filepath)
    return df[df["plan_id"] == plan_id].reset_index(drop=True)


def save_repayment_schedule(plan_id: str, records: pd.DataFrame, filepath: Path = EXCEL_FILE):
    """保存还款计划（替换该方案的所有记录）"""
    df = read_sheet(SHEET_REPAYMENT_SCHEDULE, filepath)
    df = df[df["plan_id"] != plan_id]
    df = pd.concat([df, records], ignore_index=True)
    write_sheet(df, SHEET_REPAYMENT_SCHEDULE, filepath)


def mark_period_paid(plan_id: str, period: int, pay_date: Optional[str] = None, filepath: Path = EXCEL_FILE):
    df = read_sheet(SHEET_REPAYMENT_SCHEDULE, filepath)
    mask = (df["plan_id"] == plan_id) & (df["period"] == period)
    df.loc[mask, "is_paid"] = True
    df.loc[mask, "actual_pay_date"] = pay_date or datetime.now().strftime("%Y-%m-%d")
    write_sheet(df, SHEET_REPAYMENT_SCHEDULE, filepath)


# ---- 利率调整 ----

def get_rate_adjustments(plan_id: str, filepath: Path = EXCEL_FILE) -> pd.DataFrame:
    df = read_sheet(SHEET_RATE_ADJUSTMENTS, filepath)
    return df[df["plan_id"] == plan_id].reset_index(drop=True)


def save_rate_adjustment(record: dict, filepath: Path = EXCEL_FILE):
    df = read_sheet(SHEET_RATE_ADJUSTMENTS, filepath)
    df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
    write_sheet(df, SHEET_RATE_ADJUSTMENTS, filepath)


# ---- 提前还款 ----

def get_prepayments(plan_id: str, filepath: Path = EXCEL_FILE) -> pd.DataFrame:
    df = read_sheet(SHEET_PREPAYMENTS, filepath)
    return df[df["plan_id"] == plan_id].reset_index(drop=True)


def save_prepayment(record: dict, filepath: Path = EXCEL_FILE):
    df = read_sheet(SHEET_PREPAYMENTS, filepath)
    df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
    write_sheet(df, SHEET_PREPAYMENTS, filepath)


# ---- 系统配置 ----

def get_config(key: str, filepath: Path = EXCEL_FILE) -> Optional[str]:
    df = read_sheet(SHEET_CONFIG, filepath)
    match = df[df["key"] == key]
    if match.empty:
        return None
    return str(match.iloc[0]["value"])


def get_all_config(filepath: Path = EXCEL_FILE) -> pd.DataFrame:
    """获取所有系统配置"""
    return read_sheet(SHEET_CONFIG, filepath)


def set_config(key: str, value: str, description: str = "", filepath: Path = EXCEL_FILE):
    df = read_sheet(SHEET_CONFIG, filepath)
    now = datetime.now().isoformat()
    if key in df["key"].values:
        df.loc[df["key"] == key, "value"] = value
        df.loc[df["key"] == key, "updated_at"] = now
        if description:
            df.loc[df["key"] == key, "description"] = description
    else:
        new_row = pd.DataFrame([{
            "key": key, "value": value,
            "description": description, "updated_at": now,
        }])
        df = pd.concat([df, new_row], ignore_index=True)
    write_sheet(df, SHEET_CONFIG, filepath)
